#!/usr/bin/env python3
"""Build a self-contained inspection report for the locally downloaded INVIAS archive."""

from __future__ import annotations

import hashlib
import html
import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path

import openpyxl


ARCHIVE = Path("/Users/mr/Downloads/2026_1.zip")
ISOLATED = Path("/Users/mr/Downloads/APU_8100_ARAUCA__ARAUCA_2026_1.xlsx")
OUTPUT = Path("artifacts/apu-2026-1-report.html")
SAMPLE_MEMBER = "2026_1/APU_0509_ANTIOQUIA__VALLE_DE_ABURRA_2026_1.xlsx"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def human_size(size: int) -> str:
    value = float(size)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.1f} {unit}"
        value /= 1024
    raise AssertionError("unreachable")


def display_slug(slug: str) -> str:
    special = {"SAN_ANDRES": "San Andrés", "VALLE_DEL_CAUCA": "Valle del Cauca"}
    if slug in special:
        return special[slug]
    return " ".join(word.capitalize() for word in slug.split("_"))


def main() -> None:
    if not ARCHIVE.exists() or not ISOLATED.exists():
        raise SystemExit("Expected downloads are missing")

    with zipfile.ZipFile(ARCHIVE) as archive:
        members = [info for info in archive.infolist() if info.filename.endswith(".xlsx")]
        files = []
        for info in members:
            match = re.fullmatch(
                r"2026_1/APU_(\d{4})_(.+?)__(.+?)_2026_1\.xlsx", info.filename
            )
            if not match:
                continue
            code, department_slug, province_slug = match.groups()
            files.append(
                {
                    "code": code,
                    "department": display_slug(department_slug),
                    "province": display_slug(province_slug),
                    "filename": Path(info.filename).name,
                    "compressed": info.compress_size,
                    "uncompressed": info.file_size,
                    "size": human_size(info.file_size),
                }
            )
        files.sort(key=lambda row: (row["department"], row["code"]))

        arauca_member = next(info for info in members if Path(info.filename).name == ISOLATED.name)
        isolated_bytes = ISOLATED.read_bytes()
        archive_arauca_bytes = archive.read(arauca_member)
        identical = isolated_bytes == archive_arauca_bytes

        sample_path = Path("/tmp/apu-report-sample.xlsx")
        sample_path.write_bytes(archive.read(SAMPLE_MEMBER))

    workbook = openpyxl.load_workbook(sample_path, read_only=True, data_only=True)
    index = workbook["ÍNDICE"]
    apu_sheet = workbook["APU´S"]
    apu_rows = []
    for row in range(5, 11):
        description = str(index.cell(row, 6).value or "").splitlines()[0].strip()
        apu_rows.append(
            {
                "item": index.cell(row, 5).value,
                "description": description,
                "unit": index.cell(row, 7).value,
                "equipment": index.cell(row, 8).value,
                "materials": index.cell(row, 9).value,
                "transport": index.cell(row, 10).value,
                "labor": index.cell(row, 11).value,
                "direct": index.cell(row, 12).value,
            }
        )
    apu_count = sum(
        1
        for row in range(4, apu_sheet.max_row + 1)
        if apu_sheet.cell(row, 3).value is not None
    )
    item_sheet_count = sum(bool(re.fullmatch(r"\d[\d,]*", name)) for name in workbook.sheetnames)
    sheet_names = workbook.sheetnames
    workbook.close()
    sample_path.unlink(missing_ok=True)

    department_counts = Counter(row["department"] for row in files)
    archive_hash = sha256(ARCHIVE)
    isolated_hash = sha256(ISOLATED)
    total_uncompressed = sum(row["uncompressed"] for row in files)
    source_url = "https://www.invias.gov.co/publicaciones/4149/analisis-de-precios-unitarios-apu-regionalizados-de-referencia/"

    data_json = json.dumps(files, ensure_ascii=False).replace("</", "<\\/")
    examples_json = json.dumps(apu_rows, ensure_ascii=False).replace("</", "<\\/")
    dept_chips = "".join(
        f'<span class="chip">{html.escape(dept)} <b>{count}</b></span>'
        for dept, count in sorted(department_counts.items())
    )
    sheet_chips = "".join(
        f'<span class="chip muted">{html.escape(name)}</span>' for name in sheet_names[:18]
    )

    document = f"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Radiografía del archivo APU INVIAS 2026-1</title>
  <style>
    :root{{--ink:#162033;--muted:#617086;--paper:#f5f2e9;--card:#fffdf8;--line:#d9d4c6;--yellow:#f2c94c;--blue:#265d97;--green:#2e7d60;--red:#b84d42}}
    *{{box-sizing:border-box}} body{{margin:0;background:var(--paper);color:var(--ink);font:15px/1.5 Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif}}
    a{{color:var(--blue)}} .wrap{{width:min(1180px,calc(100% - 32px));margin:auto}}
    header{{background:var(--ink);color:#fff;padding:62px 0 54px;border-bottom:8px solid var(--yellow)}}
    .eyebrow{{font-size:12px;letter-spacing:.14em;text-transform:uppercase;color:var(--yellow);font-weight:800}}
    h1{{font-size:clamp(38px,6vw,70px);line-height:.96;letter-spacing:-.045em;margin:14px 0 18px;max-width:920px}}
    header p{{font-size:18px;color:#d9e1eb;max-width:780px;margin:0}}
    .source{{display:inline-flex;margin-top:24px;padding:8px 12px;border:1px solid #667285;border-radius:999px;color:#d9e1eb;text-decoration:none}}
    main{{padding:34px 0 72px}} .grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}}
    .stat,.panel{{background:var(--card);border:1px solid var(--line);border-radius:16px;box-shadow:0 3px 0 rgba(22,32,51,.04)}}
    .stat{{padding:20px}} .stat strong{{display:block;font-size:34px;letter-spacing:-.04em}} .stat span{{color:var(--muted)}}
    .panel{{padding:26px;margin-top:18px}} .panel h2{{font-size:26px;letter-spacing:-.025em;margin:0 0 10px}}
    .verdict{{display:grid;grid-template-columns:56px 1fr;gap:18px;align-items:start;border-left:7px solid var(--yellow)}}
    .icon{{width:52px;height:52px;display:grid;place-items:center;border-radius:50%;background:#fff0b8;font-size:26px}}
    .verdict p{{font-size:17px;margin:4px 0}} .callout{{padding:14px 16px;border-radius:10px;background:#edf4fb;color:#173f68;margin-top:14px}}
    .chips{{display:flex;flex-wrap:wrap;gap:8px;margin-top:16px}} .chip{{padding:6px 10px;border-radius:999px;background:#e8efe9;color:#244a39;font-size:12px}}
    .chip b{{margin-left:4px}} .chip.muted{{background:#edf0f3;color:#47566a}}
    .toolbar{{display:grid;grid-template-columns:1fr 220px;gap:10px;margin:18px 0 12px}}
    input,select{{width:100%;border:1px solid #bfc6ce;background:#fff;padding:12px 13px;border-radius:10px;font:inherit;color:inherit}}
    .table-wrap{{overflow:auto;border:1px solid var(--line);border-radius:12px}} table{{border-collapse:collapse;width:100%;background:#fff;min-width:760px}}
    th,td{{padding:11px 13px;border-bottom:1px solid #ece8de;text-align:left}} th{{background:#f0ede5;font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:#586578;position:sticky;top:0}} td.code{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace}}
    .pill{{font-size:11px;font-weight:800;padding:4px 8px;border-radius:999px;background:#dff2e8;color:#236347}}
    .compare{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:16px}} .compare>div{{padding:16px;border:1px solid var(--line);border-radius:12px;background:#fff}}
    code{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;font-size:12px;overflow-wrap:anywhere}} .note{{color:var(--muted);font-size:13px}}
    .bar{{height:10px;background:#e7e4dc;border-radius:20px;overflow:hidden;margin-top:8px}} .bar i{{display:block;height:100%;background:var(--blue)}}
    footer{{border-top:1px solid var(--line);padding:24px 0 48px;color:var(--muted);font-size:13px}}
    @media(max-width:850px){{.grid{{grid-template-columns:1fr 1fr}}.compare{{grid-template-columns:1fr}}}}
    @media(max-width:540px){{.grid{{grid-template-columns:1fr}}.toolbar{{grid-template-columns:1fr}}header{{padding-top:42px}}.panel{{padding:19px}}}}
  </style>
</head>
<body>
<header><div class="wrap">
  <div class="eyebrow">Inspección local · periodo 2026-1</div>
  <h1>Qué contiene realmente el ZIP nacional de APU</h1>
  <p>Inventario verificable de los 140 libros provinciales, anatomía de un archivo y comparación con una descarga individual.</p>
  <a class="source" href="{source_url}">Fuente oficial: INVIAS ↗</a>
</div></header>
<main class="wrap">
  <section class="grid">
    <div class="stat"><strong>{len(files)}</strong><span>libros provinciales .xlsx</span></div>
    <div class="stat"><strong>{len(department_counts)}</strong><span>departamentos representados</span></div>
    <div class="stat"><strong>{apu_count}</strong><span>ítems APU listados en la muestra</span></div>
    <div class="stat"><strong>{human_size(total_uncompressed)}</strong><span>tamaño total sin comprimir</span></div>
  </section>

  <section class="panel verdict">
    <div class="icon">⌖</div><div><h2>Conclusión: provincia, no ciudad</h2>
    <p>El ZIP no contiene un precio distinto para cada municipio. Contiene <b>un libro por provincia DANE</b>; los municipios de una misma provincia consultan ese mismo conjunto regional de APU.</p>
    <div class="callout"><b>Ejemplo:</b> Medellín pertenece a Valle de Aburrá. Su referencia es <code>APU_0509_ANTIOQUIA__VALLE_DE_ABURRA_2026_1.xlsx</code>, compartida con los demás municipios de esa provincia. Bogotá D.C. está fuera del alcance declarado por INVIAS.</div>
    <p class="note">La relación municipio → provincia se resuelve en el filtro web de INVIAS; no aparece como una tabla municipal completa dentro del ZIP. Para URLs por ciudad, el proyecto necesitará una tabla territorial oficial separada.</p></div>
  </section>

  <section class="panel"><h2>Cobertura por departamento</h2><p>Los códigos y nombres se derivan de los nombres oficiales de archivo del ZIP.</p><div class="chips">{dept_chips}</div></section>

  <section class="panel"><h2>Explorador de los 140 archivos</h2>
    <div class="toolbar"><input id="search" aria-label="Buscar" placeholder="Buscar departamento, provincia, código…"><select id="department" aria-label="Departamento"><option value="">Todos los departamentos</option></select></div>
    <div class="table-wrap"><table><thead><tr><th>Código</th><th>Departamento</th><th>Provincia</th><th>Tamaño</th><th>Archivo</th></tr></thead><tbody id="fileRows"></tbody></table></div>
    <p class="note" id="fileCount"></p>
  </section>

  <section class="panel"><h2>Anatomía de un libro provincial</h2>
    <p>La muestra inspeccionada fue Valle de Aburrá. Tiene <b>{len(sheet_names)} hojas</b>: {item_sheet_count} hojas de ítem, además de portada, índice, tablas de insumos y hojas de cálculo.</p>
    <div class="chips">{sheet_chips}<span class="chip muted">… y {len(sheet_names)-18} más</span></div>
    <div class="callout">Cada ítem desglosa equipos, materiales, transporte y mano de obra, y entrega <b>costo directo en COP</b>. No incluye AIU y no debe presentarse como precio de mercado.</div>
    <div class="table-wrap" style="margin-top:16px"><table><thead><tr><th>Ítem</th><th>Actividad</th><th>Unidad</th><th>Equipo</th><th>Material</th><th>Transporte</th><th>Mano de obra</th><th>Costo directo</th></tr></thead><tbody id="apuRows"></tbody></table></div>
    <p class="note">Valores de ejemplo: libro oficial INVIAS Valle de Aburrá, vigencia 2026-1; COP; costos directos de referencia, sin AIU.</p>
  </section>

  <section class="panel"><h2>Prueba de descarga individual</h2>
    <p>Se descargó Arauca usando Departamento = Arauca, Municipio = Arauca, Año = 2026 y Periodo = 1. El resultado se comparó con el miembro homónimo del ZIP.</p>
    <div class="compare"><div><span class="pill">ZIP nacional</span><h3>Miembro Arauca</h3><code>{ISOLATED.name}</code><p>{human_size(arauca_member.file_size)} · {arauca_member.file_size:,} bytes</p></div>
    <div><span class="pill">Descarga individual</span><h3>Archivo aislado</h3><code>{ISOLATED.name}</code><p>{human_size(ISOLATED.stat().st_size)} · {ISOLATED.stat().st_size:,} bytes</p></div></div>
    <div class="callout"><b>{'Coincidencia exacta' if identical else 'Diferencia detectada'}:</b> {'los archivos son idénticos byte por byte.' if identical else 'los archivos no son idénticos.'}<br><code>SHA-256 {isolated_hash}</code></div>
  </section>

  <section class="panel"><h2>Proveniencia y límites de uso</h2>
    <p><b>Fuente:</b> Instituto Nacional de Vías (INVIAS), “Análisis de Precios Unitarios (APU) Regionalizados de Referencia”. <b>Vigencia:</b> 2026-1. <b>Archivo local inspeccionado:</b> <code>2026_1.zip</code>.</p>
    <p><b>SHA-256 del ZIP:</b><br><code>{archive_hash}</code></p>
    <p class="note">Los archivos oficiales permanecen fuera del repositorio. El aviso legal de la página de INVIAS restringe el uso comercial o con fines de lucro sin autorización. Este reporte describe estructura y metadatos; no redistribuye los libros fuente.</p>
  </section>
</main>
<footer><div class="wrap">Reporte generado desde archivos locales descargados de INVIAS · 2026-08-03 · America/Bogota</div></footer>
<script>
const files={data_json}; const examples={examples_json};
const tbody=document.querySelector('#fileRows'), search=document.querySelector('#search'), department=document.querySelector('#department'), count=document.querySelector('#fileCount');
[...new Set(files.map(x=>x.department))].sort((a,b)=>a.localeCompare(b,'es')).forEach(x=>department.add(new Option(x,x)));
function renderFiles(){{const q=search.value.toLocaleLowerCase('es'); const rows=files.filter(x=>(!department.value||x.department===department.value)&&(`${{x.code}} ${{x.department}} ${{x.province}} ${{x.filename}}`.toLocaleLowerCase('es').includes(q))); tbody.innerHTML=rows.map(x=>`<tr><td class="code">${{x.code}}</td><td>${{x.department}}</td><td><b>${{x.province}}</b></td><td>${{x.size}}</td><td class="code">${{x.filename}}</td></tr>`).join(''); count.textContent=`${{rows.length}} de ${{files.length}} archivos`;}}
search.addEventListener('input',renderFiles); department.addEventListener('change',renderFiles); renderFiles();
const cop=new Intl.NumberFormat('es-CO',{{style:'currency',currency:'COP',maximumFractionDigits:0}});
document.querySelector('#apuRows').innerHTML=examples.map(x=>`<tr><td class="code">${{x.item}}</td><td>${{x.description}}</td><td>${{x.unit}}</td><td>${{cop.format(x.equipment)}}</td><td>${{cop.format(x.materials)}}</td><td>${{cop.format(x.transport)}}</td><td>${{cop.format(x.labor)}}</td><td><b>${{cop.format(x.direct)}}</b></td></tr>`).join('');
</script>
</body></html>"""

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(document, encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(OUTPUT.resolve()),
                "files": len(files),
                "departments": len(department_counts),
                "apu_count": apu_count,
                "identical": identical,
                "size": OUTPUT.stat().st_size,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
